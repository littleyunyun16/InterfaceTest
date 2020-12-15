# -*- coding: utf-8 -*-
__author__ = 'amy'
__date__ = '2020/12/9 21:22'

import os
import shutil
import getpathinfo
from openpyxl import load_workbook
from openpyxl.styles import Font,Alignment
from openpyxl.styles.colors import BLACK,BLUE


path=getpathinfo.get_path()
source_file=os.path.join(path,'case','ReturnOrder.xlsx')
target_file=os.path.join(path,'result','excelReport','ReturnOrder.xlsx')


class WriteExcel():
    """文件写入数据"""
    def __init__(self,fileName):
        self.filename = fileName
        if not os.path.exists(self.filename):
            # 文件不存在，则拷贝模板文件至指定报告目录下
            shutil.copyfile(source_file,target_file)
        self.wb = load_workbook(self.filename)  # 打开一个xlsx文件
        self.ws = self.wb.active  # 激活工作表

    def write_data(self,row_n,value):
        """
        写入测试结果
        :param row_n:数据所在行数
        :param value: 测试结果值
        :return: 无
        """
        font_BLUE = Font(name='宋体', color=BLUE, bold=True)
        font_BLACk = Font(name='宋体', color=BLACK, bold=True)
        font1 = Font(name='宋体', color=BLUE, bold=True)
        align = Alignment(horizontal='center', vertical='center')
        # 获数所在行数
        # G_n = "G" + str(row_n)
        H_n = "H" + str(row_n)
        if value == "PASS":
            self.ws.cell(row_n, 8, value)
            self.ws[H_n].font = font_BLUE
        if value == "FAIL":
            self.ws.cell(row_n, 8, value)
            self.ws[H_n].font = font_BLACk
        # self.ws.cell(row_n,9, )
        self.ws[H_n].alignment = align
        # self.ws[H_n].font = font1
        # self.ws[H_n].alignment = align
        self.wb.save(self.filename)
